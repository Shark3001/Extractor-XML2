import os
import io
import xml.etree.ElementTree as ET
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
import openpyxl
from openpyxl.styles import PatternFill

app = Flask(__name__)

app.secret_key = os.getenv("SECRET_KEY_APP_XML", "CAMBIA_ESTA_CLAVE_EN_RENDER")
CORRECT_PASSWORD = os.getenv("APP_PASSWORD", "AFC2024*")

# --- Funciones de conversi칩n ---
def formatear_numero(valor):
    if valor is None:
        return ""
    # Aseguramos dos decimales para formato de presentaci칩n
    try:
        # Se asegura que si es un n칰mero, se muestre con dos decimales usando la coma como separador de decimales.
        if isinstance(valor, (int, float)):
            return f"{valor:.2f}".replace('.', ',')
        return str(valor).replace(".", ",")
    except:
        return str(valor).replace(".", ",")


def formatear_fecha(fecha_str):
    if fecha_str:
        try:
            return datetime.fromisoformat(fecha_str.replace('Z', '+00:00')).strftime('%d-%m-%Y')
        except ValueError:
            return fecha_str
    return ""

def convertir_numero(valor):
    if valor is None or valor == "":
        return 0
    
    s_valor = str(valor)
    
    # Nuevo enfoque estricto para evitar la multiplicaci칩n por error de formato:
    
    # Si detectamos una coma (',') y un punto ('.') en la cadena:
    if ',' in s_valor and '.' in s_valor:
        # Asumimos formato europeo: punto es miles, coma es decimal.
        # Ejemplo: 50.731,71  ->  50731.71
        s_valor = s_valor.replace(".", "")
        s_valor = s_valor.replace(",", ".")
    elif ',' in s_valor:
        # Asumimos formato donde solo la coma es decimal y no hay separador de miles (o es espacio).
        # Ejemplo: 50731,71 -> 50731.71
        s_valor = s_valor.replace(",", ".")
    
    # Si solo hay punto, lo dejamos como est치, asumiendo formato americano (50731.71)
    
    try:
        return float(s_valor)
    except ValueError:
        return 0

def convertir_fecha_excel(fecha_str):
    if fecha_str:
        try:
            return datetime.strptime(fecha_str, "%d-%m-%Y")
        except ValueError:
            pass
    return None

def extraer_datos_xml_en_memoria(xml_files, numero_receptor_filtro):
    wb = openpyxl.Workbook()

    # --- HOJA facturas_detalladas ---
    ws_detalladas = wb.active
    ws_detalladas.title = "facturas_detalladas"
    headers_detalladas = [
        "Clave","Consecutivo","Fecha","Nombre Emisor","N칰mero Emisor","Nombre Receptor","N칰mero Receptor",
        "C칩digo Cabys","Detalle","Cantidad","Precio Unitario","Monto Total","Monto Descuento","Subtotal",
        "Tarifa (%)","Monto Impuesto","Impuesto Neto","C칩digo Moneda","Tipo Cambio",
        "Total Gravado","Total Exento","Total Exonerado","Total Venta","Total Descuentos",
        "Total Venta Neta","Total Impuesto","Total Comprobante","Otros Cargos","Archivo","Tipo de Documento"
    ]
    ws_detalladas.append(headers_detalladas)

    # --- HOJA facturas_resumidas ---
    ws_resumidas = wb.create_sheet(title="facturas_resumidas")
    # ENCABEZADOS DE HOJA RESUMIDA
    headers_resumidas = [
        "Consecutivo",
        "Detalle",
        "Fecha",
        "C칩digo Moneda",
        "Subtotal",
        "Total Descuentos",
        "Total Impuesto",
        "Otros Cargos",
        "N칰mero Receptor",
        "Total Comprobante" 
    ]
    ws_resumidas.append(headers_resumidas)

    # --- HOJA facturas_resumidasV2 (NUEVA HOJA) ---
    ws_resumidas_v2 = wb.create_sheet(title="facturas_resumidasV2")
    # ENCABEZADOS DE HOJA RESUMIDA V2
    headers_resumidas_v2 = [
        "Consecutivo",
        "Fecha",
        "Nombre Emisor",
        "N칰mero Emisor",
        "Nombre Receptor",
        "N칰mero Receptor",
        "Tarifa (%)", # Se tomar치 la tarifa de la primera l칤nea de detalle o un promedio si se necesita, pero generalmente se usa el campo de resumen si existe. Como no existe en resumen, usaremos la tarifa de la primera l칤nea (como proxy).
        "Total Descuentos",
        "Total Venta Neta",
        "Total Impuesto",
        "Total Comprobante",
        "Otros Cargos",
        "Archivo",
        "Tipo de Documento"
    ]
    ws_resumidas_v2.append(headers_resumidas_v2)


    for uploaded_file in xml_files:
        filename = uploaded_file.filename
        try:
            tree = ET.parse(uploaded_file)
            root = tree.getroot()
            for elem in root.iter():
                elem.tag = elem.tag.split('}', 1)[-1]

            tipo_documento = root.tag.split(' ')[0]
            if tipo_documento == "MensajeHacienda":
                continue

            clave = root.find('Clave').text if root.find('Clave') is not None else ""
            consecutivo = root.find('NumeroConsecutivo').text if root.find('NumeroConsecutivo') is not None else ""
            fecha = formatear_fecha(root.find('FechaEmision').text) if root.find('FechaEmision') is not None else ""
            nombre_emisor = root.find('Emisor/Nombre').text if root.find('Emisor/Nombre') is not None else ""
            numero_emisor = root.find('Emisor/Identificacion/Numero').text if root.find('Emisor/Identificacion/Numero') is not None else ""
            nombre_receptor = root.find('Receptor/Nombre').text if root.find('Receptor/Nombre') is not None else ""
            numero_receptor = root.find('Receptor/Identificacion/Numero').text if root.find('Receptor/Identificacion/Numero') is not None else ""
            
            # 游늷 Obtener fecha corta (dd-mm-yy) para la concatenaci칩n
            fecha_dd_mm_yy = ""
            if fecha and fecha != "":
                try:
                    dt_object = datetime.strptime(fecha, '%d-%m-%Y') 
                    fecha_dd_mm_yy = dt_object.strftime('%d-%m-%y')
                except ValueError:
                    fecha_dd_mm_yy = fecha 

            resumen_factura = root.find('ResumenFactura')
            total_venta = formatear_numero(resumen_factura.find('TotalVenta').text) if resumen_factura is not None and resumen_factura.find('TotalVenta') is not None else ""
            total_descuentos = formatear_numero(resumen_factura.find('TotalDescuentos').text) if resumen_factura is not None and resumen_factura.find('TotalDescuentos') is not None else ""
            total_venta_neta = formatear_numero(resumen_factura.find('TotalVentaNeta').text) if resumen_factura is not None and resumen_factura.find('TotalVentaNeta') is not None else ""
            total_exento = formatear_numero(resumen_factura.find('TotalExento').text) if resumen_factura is not None and resumen_factura.find('TotalExento') is not None else ""
            total_exonerado = formatear_numero(resumen_factura.find('TotalExonerado').text) if resumen_factura is not None and resumen_factura.find('TotalExonerado') is not None else ""
            total_impuesto = formatear_numero(resumen_factura.find('TotalImpuesto').text) if resumen_factura is not None and resumen_factura.find('TotalImpuesto') is not None else ""
            total_comprobante = formatear_numero(resumen_factura.find('TotalComprobante').text) if resumen_factura is not None and resumen_factura.find('TotalComprobante') is not None else ""
            otros_cargos = formatear_numero(root.find('OtrosCargos/MontoCargo').text) if root.find('OtrosCargos/MontoCargo') is not None else "0,00"
            codigo_moneda = root.find('ResumenFactura/CodigoTipoMoneda/CodigoMoneda').text if root.find('ResumenFactura/CodigoTipoMoneda/CodigoMoneda') is not None else ""
            
            detalles_servicio = root.find('DetalleServicio')
            detalle_texto = "" # Cadena final concatenada para facturas_resumidas
            subtotal_factura = 0 
            tarifa_resumen = "0,00" # Tarifa para facturas_resumidasV2

            if detalles_servicio is not None:
                lineas_detalle = detalles_servicio.findall('LineaDetalle')
                
                # L칍GICA DE CONCATENACI칍N DE DETALLE DE L칈NEAS EXISTENTE
                detalle_texto_lineas = "; ".join([linea.find('Detalle').text if linea.find('Detalle') is not None else "" for linea in lineas_detalle])
                
                # CONCATENACI칍N FINAL REQUERIDA: Fecha corta + Nombre Emisor + Detalles de l칤neas
                detalle_texto = f"{fecha_dd_mm_yy} - {nombre_emisor} - {detalle_texto_lineas}"
                
                # C츼LCULO DEL SUBTOTAL: Suma de SubTotales de l칤neas 
                for linea in lineas_detalle:
                    subtotal_linea_str = linea.find('SubTotal').text if linea.find('SubTotal') is not None else "0"
                    subtotal_factura += convertir_numero(subtotal_linea_str)
                    
                    # Obtener la tarifa de la primera l칤nea como proxy para Tarifa (%) en V2
                    if tarifa_resumen == "0,00":
                        impuesto = linea.find('Impuesto')
                        tarifa_resumen = formatear_numero(impuesto.find('Tarifa').text) if impuesto is not None and impuesto.find('Tarifa') is not None else "0,00"
            else:
                # CONCATENACI칍N FINAL si no hay detalles de l칤nea
                detalle_texto = f"{fecha_dd_mm_yy} - {nombre_emisor} - (Sin detalles)"


            # --- facturas_detalladas ---
            if detalles_servicio is not None:
                for linea in detalles_servicio.findall('LineaDetalle'):
                    codigo_cabys = linea.find('Codigo').text if linea.find('Codigo') is not None else ""
                    detalle = linea.find('Detalle').text if linea.find('Detalle') is not None else ""
                    cantidad = formatear_numero(linea.find('Cantidad').text) if linea.find('Cantidad') is not None else ""
                    precio_unitario = formatear_numero(linea.find('PrecioUnitario').text) if linea.find('PrecioUnitario') is not None else ""
                    monto_total_linea = formatear_numero(linea.find('MontoTotal').text) if linea.find('MontoTotal') is not None else ""
                    monto_descuento_linea = formatear_numero(linea.find('Descuento/MontoDescuento').text) if linea.find('Descuento/MontoDescuento') is not None else "0,00"
                    subtotal_linea = formatear_numero(linea.find('SubTotal').text) if linea.find('SubTotal') is not None else ""
                    impuesto = linea.find('Impuesto')
                    tarifa_linea = formatear_numero(impuesto.find('Tarifa').text) if impuesto is not None and impuesto.find('Tarifa') is not None else "0,00"
                    monto_impuesto_linea = formatear_numero(impuesto.find('Monto').text) if impuesto is not None and impuesto.find('Monto') is not None else "0,00"
                    impuesto_neto_linea = formatear_numero(linea.find('ImpuestoNeto').text) if linea.find('ImpuestoNeto') is not None else ""
                    codigo_moneda_linea = root.find('ResumenFactura/CodigoTipoMoneda/CodigoMoneda').text if root.find('ResumenFactura/CodigoTipoMoneda/CodigoMoneda') is not None else ""
                    tipo_cambio = formatear_numero(root.find('ResumenFactura/CodigoTipoMoneda/TipoCambio').text) if root.find('ResumenFactura/CodigoTipoMoneda/TipoCambio') is not None else ""
                    total_gravado = formatear_numero(root.find('ResumenFactura/TotalGravado').text) if root.find('ResumenFactura/TotalGravado') is not None else ""
                    total_comprobante_linea = total_comprobante
                    otros_cargos_linea = otros_cargos

                    fila_detallada = [
                        clave,
                        consecutivo,
                        convertir_fecha_excel(fecha),
                        nombre_emisor,
                        numero_emisor,
                        nombre_receptor,
                        numero_receptor,
                        codigo_cabys,
                        detalle,
                        convertir_numero(cantidad),
                        convertir_numero(precio_unitario),
                        convertir_numero(monto_total_linea),
                        convertir_numero(monto_descuento_linea),
                        convertir_numero(subtotal_linea),
                        convertir_numero(tarifa_linea),
                        convertir_numero(monto_impuesto_linea),
                        convertir_numero(impuesto_neto_linea),
                        codigo_moneda_linea,
                        convertir_numero(tipo_cambio),
                        convertir_numero(total_gravado),
                        convertir_numero(total_exento),
                        convertir_numero(total_exonerado),
                        convertir_numero(total_venta),
                        convertir_numero(total_descuentos),
                        convertir_numero(total_venta_neta),
                        convertir_numero(total_impuesto),
                        convertir_numero(total_comprobante_linea),
                        convertir_numero(otros_cargos_linea),
                        filename,
                        tipo_documento
                    ]
                    ws_detalladas.append(fila_detallada)

            # --- facturas_resumidas ---
            fila_resumida = [
                consecutivo,
                detalle_texto, 
                convertir_fecha_excel(fecha),
                codigo_moneda,
                subtotal_factura, 
                convertir_numero(total_descuentos),
                convertir_numero(total_impuesto),
                convertir_numero(otros_cargos),
                numero_receptor,
                convertir_numero(total_comprobante) 
            ]
            ws_resumidas.append(fila_resumida)
            
            # --- facturas_resumidasV2 (NUEVA FILA) ---
            fila_resumida_v2 = [
                consecutivo,
                convertir_fecha_excel(fecha),
                nombre_emisor,
                numero_emisor,
                nombre_receptor,
                numero_receptor,
                convertir_numero(tarifa_resumen), # Usamos el valor proxy de la primera l칤nea
                convertir_numero(total_descuentos),
                convertir_numero(total_venta_neta),
                convertir_numero(total_impuesto),
                convertir_numero(total_comprobante),
                convertir_numero(otros_cargos),
                filename,
                tipo_documento
            ]
            ws_resumidas_v2.append(fila_resumida_v2)


        except Exception as e:
            flash(f"Error al procesar '{filename}': {e}", 'error')

    # --- Formato colores facturas_detalladas ---
    fill_celeste = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    fill_rojo = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
    # Columnas azules en hoja detallada: B, C, I, P, V, X, Y, AC
    col_azul = ["B","C","I","P","V","X","Y","AC"]
    for col in col_azul:
        for cell in ws_detalladas[col]:
            cell.fill = fill_celeste
    # Columna roja en hoja detallada (N칰mero Receptor): G
    for cell in ws_detalladas["G"][1:]:
        if cell.value and numero_receptor_filtro and str(cell.value) != str(numero_receptor_filtro):
            cell.fill = fill_rojo

    # --- Formato colores facturas_resumidas ---
    for fila in ws_resumidas.iter_rows(min_row=2):
        cell_receptor = fila[8] # Columna 9 (칈ndice 8)
        
        for i, cell in enumerate(fila):
            if i != 8:
                cell.fill = PatternFill(fill_type=None)
            
        if cell_receptor.value and numero_receptor_filtro and str(cell_receptor.value) != str(numero_receptor_filtro):
            cell_receptor.fill = fill_rojo
        else:
            cell_receptor.fill = PatternFill(fill_type=None)
        
        # APLICACI칍N DE FORMATO NUM칄RICO EXPLICITO
        # 칈ndices de columnas de monto (0-based): 4: Subtotal, 5: T. Descuentos, 6: T. Impuesto, 7: Otros Cargos, 9: T. Comprobante
        column_indices_to_format = [4, 5, 6, 7, 9] 
        for col_index_to_format in column_indices_to_format:
            cell_to_format = fila[col_index_to_format]
            if isinstance(cell_to_format.value, (int, float)):
                cell_to_format.number_format = '#,##0.00' 
                
    # --- Formato colores facturas_resumidasV2 (NUEVO FORMATO) ---
    for fila in ws_resumidas_v2.iter_rows(min_row=2):
        cell_receptor_v2 = fila[5] # Columna 6 (N칰mero Receptor, 칈ndice 5)
        cell_tarifa_v2 = fila[6] # Columna 7 (Tarifa (%), 칈ndice 6)
        
        # Eliminamos relleno de todas las celdas (deber칤an estar blancas)
        for i, cell in enumerate(fila):
             cell.fill = PatternFill(fill_type=None)
            
        # Aplicamos el color rojo (si aplica) al N칰mero Receptor
        if cell_receptor_v2.value and numero_receptor_filtro and str(cell_receptor_v2.value) != str(numero_receptor_filtro):
            cell_receptor_v2.fill = fill_rojo
        
        # APLICACI칍N DE FORMATO NUM칄RICO EXPLICITO
        # 칈ndices de columnas de monto (0-based) en V2:
        # 6: Tarifa (%), 7: T. Descuentos, 8: T. Venta Neta, 9: T. Impuesto, 10: T. Comprobante, 11: Otros Cargos
        column_indices_to_format_v2 = [6, 7, 8, 9, 10, 11] 
        for col_index_to_format in column_indices_to_format_v2:
            cell_to_format = fila[col_index_to_format]
            if isinstance(cell_to_format.value, (int, float)):
                # La tarifa lleva formato de porcentaje con 2 decimales
                if col_index_to_format == 6:
                    cell_to_format.number_format = '0.00%'
                else:
                    cell_to_format.number_format = '#,##0.00' 


    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    # Limpiar archivos en memoria
    xml_files.clear()
    return out

# --------- Rutas ---------
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == CORRECT_PASSWORD:
            session['logged_in'] = True
            flash('Inicio de sesi칩n exitoso.', 'success')
            return redirect(url_for('index'))
        else:
            flash('Contrase침a incorrecta. Int칠ntalo de nuevo.', 'error')
            return redirect(url_for('login'))
    return render_template('login.html')

@app.route('/')
def index():
    if not session.get('logged_in'):
        return redirect(url_for('login'))
    return render_template('index.html')

@app.route('/logout')
def logout():
    session.pop('logged_in', None)
    flash('Has cerrado sesi칩n correctamente.', 'success')
    return redirect(url_for('login'))

@app.route('/upload', methods=['POST'])
def upload_files():
    if not session.get('logged_in'):
        flash('Por favor, inicia sesi칩n para acceder a esta funci칩n.', 'error')
        return redirect(url_for('login'))

    if 'xml_files' not in request.files:
        flash('No se subieron archivos.', 'error')
        return redirect(url_for('index'))

    files = request.files.getlist('xml_files')
    if not files or files[0].filename == '':
        flash('No se seleccion칩 ning칰n archivo.', 'error')
        return redirect(url_for('index'))

    numero_receptor = request.form.get('numero_receptor')
    if not numero_receptor:
        flash('El n칰mero de identificaci칩n del receptor es obligatorio.', 'error')
        return redirect(url_for('index'))

    excel_stream = extraer_datos_xml_en_memoria(files, numero_receptor)

    return send_file(
        excel_stream,
        download_name='datos_facturas.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=False)
